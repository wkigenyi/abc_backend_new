from django.http import Http404
from django.shortcuts import render

from django.shortcuts import render
from rest_framework import generics,viewsets
from recon.db_exceptions import select_exceptions
from recon.db_recon_stats import recon_stats_req
from io import BytesIO
from zipfile import ZipFile
from django.http import FileResponse, Http404

from recon.db_reversals import select_reversals
from recon.mainFile import reconcileMain
from recon.setle_sabs import setleSabs, unserializable_floats
from recon.setlement_ import settle
from .models import Recon,ReconLog,UploadedFile
from .serializers import ReconcileSerializer, ReconciliationSerializer, SabsSerializer, SettlementSerializer,UploadedFileSerializer,LogSerializer,TransactionSerializer
from openpyxl import load_workbook
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import generics,status
from .models import Bank,UserBankMapping,Transactions
from django.db.models import Q
from dotenv import load_dotenv
import os
# Load the .env file
load_dotenv()

# Get the environment variables
server = os.getenv('DB_SERVER')
database = os.getenv('DB_NAME')
username = os.getenv('DB_USERNAME')
password = os.getenv('DB_PASSWORD')

# Create your views here.

def get_swift_code_from_request(request):
    user = request.user
    mapping = UserBankMapping.objects.filter(user=user)[0]
    bank = mapping.bank
    swift_code = bank.swift_code
    return swift_code

def get_bank_code_from_request(request):
    user = request.user
    mapping = UserBankMapping.objects.filter(user=user)[0]
    bank = mapping.bank
    bank_code = bank.bank_code
    return bank_code

class ReconciliationListView(generics.ListCreateAPIView):
    queryset = Recon.objects.all()
    serializer_class = ReconciliationSerializer

class ReconciliationLogListView(generics.ListAPIView):
    queryset = ReconLog.objects.all()

def upload_reconciliations(request):
    pass

class UploadedFilesViewset(viewsets.ModelViewSet):
    queryset = UploadedFile.objects.all()
    serializer_class = UploadedFileSerializer
    def create(self, request, *args, **kwargs):
        user = request.user
        file = request.FILES['file']
        wb = load_workbook(file)
        sheet = wb.get_sheet_by_name("Sheet1")
        start_row = 2
        count = 0
        for _ in sheet.iter_rows(min_row=start_row,max_row=10):
            row_no = start_row+count
            time = sheet.cell(row_no,1).value
            transaction_type = sheet.cell(row_no,2).value
            amount = sheet.cell(row_no,3).value
            abc_reference = sheet.cell(row_no,4).value
            recon = Recon(
                date_time=time,
                last_modified_by_user=user,
                trn_ref=abc_reference,
                )
            recon.save()
            print(time,transaction_type,amount,abc_reference)
            count+=1
        
        return super().create(request, *args, **kwargs) 

class ReconcileView(APIView):
    serializer_class = ReconcileSerializer
    # permission_classes = [IsAuthenticated]


    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        user = request.user
        if serializer.is_valid():
            uploaded_file = serializer.validated_data['file']
            bank_code = get_bank_code_from_request(request)

            

            # Save the uploaded file temporarily
            temp_file_path = "temp_file.xlsx"
            with open(temp_file_path, "wb") as buffer:
                buffer.write(uploaded_file.read())

            try:
                # Call the main function with the path of the saved file and the swift code
                merged_df, reconciled_data, succunreconciled_data, exceptions, feedback, requestedRows, UploadedRows, date_range_str = reconcileMain(
                    temp_file_path, bank_code,user)
                
                # Perform clean up: remove the temporary file after processing
                os.remove(temp_file_path)
                
                data = {
                    "reconciledRows": len(reconciled_data),
                    "unreconciledRows": len(succunreconciled_data),
                    "exceptionsRows": len(exceptions),
                    "feedback": feedback,
                    "RequestedRows": requestedRows,
                    "UploadedRows": UploadedRows,
                    "min_max_DateRange": date_range_str
                }

                return Response(data, status=status.HTTP_200_OK)

            except Exception as e:
                # If there's an error during the process, ensure the temp file is removed
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                
                # Return error as response
                return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

class ReversalsView(generics.ListAPIView):
    serializer_class = TransactionSerializer
    """
    Retrieve reversal data.
    """

    def get_queryset(self):
        # Use values from .env for database connection
        bank_code = get_bank_code_from_request(self.request)
        return Transactions.objects.filter(Q(request_type="1200")& (Q(issuer_code = bank_code)|Q(acquirer_code = bank_code))).exclude(amount="0").exclude(trn_status_0=None).exclude(trn_status_1=None)
        


        # The data returned by select_reversals is assumed to be in a suitable format for JSON serialization

        return Response(data, status=status.HTTP_200_OK)
class ExceptionsView(generics.ListAPIView):
    serializer_class = ReconciliationSerializer
    def get_queryset(self):
        bank_code = get_bank_code_from_request(self.request)
        return Recon.objects.filter(Q(issuer_code = bank_code)|Q(acquirer_code=bank_code)).exclude(excep_flag = None)
    
    
class ReconciledDataView(APIView):
    """
    Retrieve reconciled data.
    """

    def get(self, request, *args, **kwargs):
        global reconciled_data

        if reconciled_data is not None:
            reconciled_data_cleaned = unserializable_floats(reconciled_data)
            data = reconciled_data_cleaned.to_dict(orient='records')
            return Response(data, status=status.HTTP_200_OK)
        else:
            raise Http404("Reconciled data not found")
        
class UnReconciledDataView(APIView):
    """
    Retrieve unreconciled data.
    """

    def get(self, request, *args, **kwargs):
        global succunreconciled_data

        if succunreconciled_data is not None:
            
            # reconciled_data_cleaned = unserializable_floats(reconciled_data)
            # data = reconciled_data_cleaned.to_dict(orient='records')

            unreconciled_data_cleaned = unserializable_floats(succunreconciled_data)
            data =  unreconciled_data_cleaned.to_dict(orient='records')
            return Response(data, status=status.HTTP_200_OK)
        else:
            raise Http404("unReconciled data not found")  

class ReconStatsView(generics.ListAPIView):
    serializer_class = LogSerializer
    def get_queryset(self):
        bank_code = get_bank_code_from_request(self.request)
        return ReconLog.objects.filter(bank_id=bank_code)
    
    """ def get(self, request):
            
            try:
                # Assume recon_stats_req returns a list of dictionaries or None
                data = recon_stats_req(bank_code)
                if data is None:
                    return Response({'error': 'No data found'}, status=status.HTTP_404_NOT_FOUND)              
                
                return Response(data, status=status.HTTP_200_OK)

            except Exception as e:
                return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR) """

class sabsreconcile_csv_filesView(APIView):

    serializer_class = SabsSerializer

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid():
            uploaded_file = serializer.validated_data['file']
            batch_number = serializer.validated_data['batch_number']

            # Save the uploaded file temporarily
            temp_file_path = "temp_file.xlsx"
            with open(temp_file_path, "wb") as buffer:
                buffer.write(uploaded_file.read())

            try:
                # Assume setleSabs returns dataframes as one of its outputs
                _, matched_setle, _, unmatched_setlesabs = setleSabs(temp_file_path, batch_number)
                
                # Perform clean up: remove the temporary file after processing
                os.remove(temp_file_path)

                matched_csv = matched_setle.to_csv(index=False)
                unmatched_csv = unmatched_setlesabs.to_csv(index=False)

                # Create a zip file in memory
                memory_file = BytesIO()
                with ZipFile(memory_file, 'w') as zf:
                    zf.writestr('matched_setle.csv', matched_csv)
                    zf.writestr('unmatched_setlesabs.csv', unmatched_csv)
                
                # u will figure ou how to retun a zipped file here
                memory_file.seek(0)

                response = FileResponse(memory_file, content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename=Settlement_.zip'
                return response
            
            # this one below doesnt return a zipped file. gets a decode error
                # memory_file.seek(0)

                # response = Response(memory_file, content_type='application/zip')
                # response['Content-Disposition'] = 'attachment; filename=Settlement_recon_files.zip'
                # return response

            except Exception as e:
                # If there's an error during the process, ensure the temp file is removed
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                
                # Return error as response
                return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class SettlementView(APIView):
    serializer_class = SettlementSerializer

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid():
            batch_number = serializer.validated_data['batch_number']

            try:
                # Assume the settle function is defined and available here
                settlement_result = settle(batch_number)

                # Handle case where no records were found or an error occurred in settle
                if settlement_result is None or settlement_result.empty:
                    return Response({"detail": "No records for processing found or an error occurred."},
                                    status=status.HTTP_400_BAD_REQUEST)

                # Convert the DataFrame to CSV
                settlement_csv = settlement_result.to_csv(index=False)

                # Create a zip file in memory
                memory_file = BytesIO()
                with ZipFile(memory_file, 'w') as zf:
                    zf.writestr('settlement_result.csv', settlement_csv)

                memory_file.seek(0)

                response = FileResponse(memory_file, content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename=Settlement_.zip'
                return response

            except Exception as e:
                # Handle other unexpected errors
                return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

